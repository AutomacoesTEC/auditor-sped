"""Gerador de relatorio Excel (XLSX) para o Auditor SPED.

Estrutura de abas:
  - Capa: identificacao da empresa, periodo, totais
  - Resumo: todos os achados consolidados por tributo
  - R01..R10: detalhamento por regra
  - Gatilhos_Conta: gatilhos de conta (alta e media severidade)
  - Gatilhos_Hist: amostra dos gatilhos de historico

Uso:
  from src.relatorios.gerador_xlsx import gerar_xlsx
  gerar_xlsx(resultado_dict, caminho_saida)
"""

import logging
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _openpyxl_ok = True
except ImportError:
    openpyxl = None
    _openpyxl_ok = False

logger = logging.getLogger(__name__)

# ============================================================
# PALETA DE CORES (strings hexadecimais - sem dependencia de lib)
# ============================================================
_COR_TITULO = "1F3864"
_COR_CABEC = "2E75B6"
_COR_ALTA = "C00000"
_COR_MEDIA = "ED7D31"
_COR_BAIXA = "70AD47"
_COR_ALTERNADA = "DEEAF1"
_COR_BRANCO = "FFFFFF"
_COR_FUNDO_CAPA = "1F3864"

_FORMATO_MONETARIO = 'R$ #,##0.00'


def _borda_fina():
    return Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def gerar_xlsx(resultado: dict, caminho_saida: str) -> str:
    """
    Gera arquivo Excel com o relatorio completo de auditoria.

    Args:
        resultado: dicionario com dados da auditoria (formato main.py)
        caminho_saida: caminho completo do arquivo .xlsx a gerar

    Returns:
        Caminho do arquivo gerado.
    """
    if not _openpyxl_ok:
        logger.error("openpyxl nao instalado. Execute: pip install openpyxl")
        raise ImportError("openpyxl e necessario para gerar o relatorio Excel.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove aba padrao

    empresa = resultado.get("empresa", "")
    cnpj = resultado.get("cnpj", "")
    periodo = resultado.get("periodo", "")
    achados = resultado.get("regras", {}).get("achados", [])
    valor_total = resultado.get("regras", {}).get("valor_estimado", 0.0)
    gatilhos = resultado.get("gatilhos", {}).get("achados", [])
    resumo_gat = resultado.get("gatilhos", {}).get("resumo", {})

    _criar_aba_capa(wb, empresa, cnpj, periodo, achados, valor_total, resumo_gat)
    _criar_aba_resumo(wb, achados, valor_total)
    _criar_abas_por_regra(wb, achados)
    _criar_aba_gatilhos_conta(wb, gatilhos)
    _criar_aba_gatilhos_historico(wb, gatilhos)

    wb.save(caminho_saida)
    logger.info("Relatorio Excel salvo em: %s", caminho_saida)
    return caminho_saida


# ============================================================
# ABA CAPA
# ============================================================

def _criar_aba_capa(wb, empresa, cnpj, periodo, achados, valor_total, resumo_gat):
    ws = wb.create_sheet("Capa")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50

    # Cabecalho
    _mesclar_titulo(ws, "A1:B1", "RELATORIO DE AUDITORIA TRIBUTARIA - SPED", _COR_FUNDO_CAPA, 14)
    _mesclar_titulo(ws, "A2:B2", "Sistema de Auditoria de Creditos Tributarios (Lucro Real)", _COR_CABEC, 11)

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 22

    # Dados da empresa
    dados = [
        ("", ""),
        ("EMPRESA:", empresa),
        ("CNPJ:", cnpj),
        ("PERIODO:", periodo),
        ("DATA DO RELATORIO:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
    ]
    linha = 3
    for rot, val in dados:
        ws.cell(linha, 1, rot).font = Font(bold=True)
        ws.cell(linha, 2, val)
        linha += 1

    # Resumo financeiro
    _mesclar_titulo(ws, f"A{linha}:B{linha}", "RESUMO DOS ACHADOS", _COR_CABEC)
    linha += 1

    total_por_tributo: dict[str, float] = {}
    total_por_conf: dict[str, int] = {"alta": 0, "media": 0, "baixa": 0}
    for a in achados:
        trib = a.get("tributo", "Outros")
        total_por_tributo[trib] = total_por_tributo.get(trib, 0) + a.get("valor_estimado", 0)
        conf = a.get("confianca", "baixa")
        total_por_conf[conf] = total_por_conf.get(conf, 0) + 1

    resumo_dados = [
        ("Total de achados por regras:", len(achados)),
        ("Valor total estimado (creditos):", f"R$ {valor_total:,.2f}"),
        ("Achados de alta confianca:", total_por_conf["alta"]),
        ("Achados de media confianca:", total_por_conf["media"]),
        ("Achados de baixa confianca:", total_por_conf["baixa"]),
        ("", ""),
    ]
    for rot, val in resumo_dados:
        c1 = ws.cell(linha, 1, rot)
        c2 = ws.cell(linha, 2, str(val))
        c1.font = Font(bold=True)
        if "Valor total" in rot:
            c2.font = Font(bold=True, color=_COR_ALTA, size=12)
        linha += 1

    # Por tributo
    _mesclar_titulo(ws, f"A{linha}:B{linha}", "VALOR ESTIMADO POR TRIBUTO", _COR_CABEC)
    linha += 1
    for trib, val in sorted(total_por_tributo.items(), key=lambda x: -x[1]):
        ws.cell(linha, 1, trib).font = Font(bold=True)
        c = ws.cell(linha, 2, val)
        c.number_format = _FORMATO_MONETARIO
        c.font = Font(bold=True)
        linha += 1

    # Gatilhos
    linha += 1
    _mesclar_titulo(ws, f"A{linha}:B{linha}", "RESUMO DOS GATILHOS", _COR_CABEC)
    linha += 1
    por_sev = resumo_gat.get("por_severidade", {})
    gat_dados = [
        ("Total de gatilhos disparados:", resumo_gat.get("total", 0)),
        ("Alta severidade:", por_sev.get("alta", 0)),
        ("Media severidade:", por_sev.get("media", 0)),
        ("Baixa severidade:", por_sev.get("baixa", 0)),
    ]
    for rot, val in gat_dados:
        ws.cell(linha, 1, rot).font = Font(bold=True)
        ws.cell(linha, 2, str(val))
        linha += 1

    # Aviso legal
    linha += 2
    _mesclar_titulo(
        ws,
        f"A{linha}:B{linha}",
        "AVISO: Este relatorio tem carater preliminar. Achados de media/baixa confianca "
        "requerem analise documental antes de qualquer acao fiscal.",
        "FF0000",
        8,
    )

    ws.sheet_view.showGridLines = False


# ============================================================
# ABA RESUMO
# ============================================================

def _criar_aba_resumo(wb, achados, valor_total):
    ws = wb.create_sheet("Resumo")

    cabecalhos = [
        "Regra", "Titulo", "Tributo", "Valor Estimado (R$)",
        "Confianca", "Base Legal", "Recomendacao"
    ]
    larguras = [10, 55, 12, 22, 12, 40, 55]
    _cabecalho_tabela(ws, 1, cabecalhos, larguras)

    for i, a in enumerate(achados, start=2):
        linha_par = i % 2 == 0
        cor_fundo = _COR_ALTERNADA if linha_par else _COR_BRANCO
        conf = a.get("confianca", "baixa")
        cor_conf = _COR_ALTA if conf == "alta" else (_COR_MEDIA if conf == "media" else _COR_BAIXA)

        _celula(ws, i, 1, a.get("regra", ""), cor_fundo, negrito=True)
        _celula(ws, i, 2, a.get("titulo", ""), cor_fundo)
        _celula(ws, i, 3, a.get("tributo", ""), cor_fundo, centralizado=True)
        c = ws.cell(i, 4, a.get("valor_estimado", 0))
        c.number_format = _FORMATO_MONETARIO
        c.fill = PatternFill("solid", fgColor=cor_fundo)
        c.alignment = Alignment(horizontal="right")
        c.border = _borda_fina()
        _celula(ws, i, 5, conf.upper(), cor_conf, cor_texto="FFFFFF", centralizado=True, negrito=True)
        _celula(ws, i, 6, a.get("base_legal", "")[:60], cor_fundo)
        _celula(ws, i, 7, a.get("recomendacao", "")[:200], cor_fundo, wrap=True)
        ws.row_dimensions[i].height = 40

    # Linha de total
    linha_total = len(achados) + 2
    _celula(ws, linha_total, 1, "TOTAL", _COR_TITULO, cor_texto="FFFFFF", negrito=True, span=3)
    c = ws.cell(linha_total, 4, valor_total)
    c.number_format = _FORMATO_MONETARIO
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=_COR_TITULO)
    c.border = _borda_fina()
    c.alignment = Alignment(horizontal="right")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


# ============================================================
# ABAS POR REGRA
# ============================================================

def _criar_abas_por_regra(wb, achados):
    por_regra: dict[str, list] = {}
    for a in achados:
        regra = a.get("regra", "?")
        por_regra.setdefault(regra, []).append(a)

    for regra, lista in sorted(por_regra.items()):
        nome_aba = regra  # Ex: "R01"
        ws = wb.create_sheet(nome_aba)

        # Titulo
        titulo = lista[0].get("titulo", regra) if lista else regra
        _mesclar_titulo(ws, "A1:B1", f"{regra} - {titulo[:80]}", _COR_CABEC, 11)
        ws.row_dimensions[1].height = 25

        linha = 3
        for achado in lista:
            _mesclar_titulo(ws, f"A{linha}:B{linha}", achado.get("titulo", ""), _COR_TITULO, 10)
            linha += 1

            campos = [
                ("Tributo:", achado.get("tributo", "")),
                ("Valor Estimado:", f"R$ {achado.get('valor_estimado', 0):,.2f}"),
                ("Confianca:", achado.get("confianca", "").upper()),
                ("Base Legal:", achado.get("base_legal", "")),
                ("", ""),
                ("Descricao:", achado.get("descricao", "")),
                ("", ""),
                ("Recomendacao:", achado.get("recomendacao", "")),
                ("Risco:", achado.get("risco", "")),
                ("", ""),
            ]
            for rot, val in campos:
                c1 = ws.cell(linha, 1, rot)
                c2 = ws.cell(linha, 2, val)
                c1.font = Font(bold=True)
                if rot == "Valor Estimado:":
                    c2.font = Font(bold=True, color=_COR_ALTA)
                c2.alignment = Alignment(wrap_text=True)
                ws.row_dimensions[linha].height = max(15, len(str(val)) // 60 * 15 + 15)
                linha += 1

            # Registros de origem
            registros = achado.get("registros_origem", [])
            if registros:
                _mesclar_titulo(ws, f"A{linha}:B{linha}", "Registros de Origem", _COR_CABEC, 10)
                linha += 1
                for reg in registros:
                    ws.cell(linha, 1, "")
                    ws.cell(linha, 2, reg).alignment = Alignment(wrap_text=True)
                    ws.row_dimensions[linha].height = 18
                    linha += 1

            linha += 1  # Espaco entre achados

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 90
        ws.sheet_view.showGridLines = False


# ============================================================
# ABA GATILHOS DE CONTA
# ============================================================

def _criar_aba_gatilhos_conta(wb, gatilhos):
    ws = wb.create_sheet("Gatilhos_Conta")
    contas = [g for g in gatilhos if g.get("tipo") == "conta"]

    cabecalhos = ["Gatilho", "Severidade", "Categoria", "Conta", "Saldo (R$)", "Acao Recomendada"]
    larguras = [10, 12, 20, 40, 18, 60]
    _cabecalho_tabela(ws, 1, cabecalhos, larguras)

    # Ordena por severidade (alta primeiro)
    ordem_sev = {"alta": 0, "media": 1, "baixa": 2}
    contas_ord = sorted(contas, key=lambda x: (ordem_sev.get(x.get("severidade", "baixa"), 3), x.get("gatilho", "")))

    for i, g in enumerate(contas_ord, start=2):
        sev = g.get("severidade", "baixa")
        cor_sev = _COR_ALTA if sev == "alta" else (_COR_MEDIA if sev == "media" else _COR_BAIXA)
        linha_par = i % 2 == 0
        cor_fundo = _COR_ALTERNADA if linha_par else _COR_BRANCO

        _celula(ws, i, 1, g.get("gatilho", ""), cor_fundo, negrito=True, centralizado=True)
        _celula(ws, i, 2, sev.upper(), cor_sev, cor_texto="FFFFFF", centralizado=True, negrito=True)
        _celula(ws, i, 3, g.get("categoria", ""), cor_fundo, centralizado=True)
        _celula(ws, i, 4, g.get("conta", ""), cor_fundo)
        c = ws.cell(i, 5, g.get("valor", 0))
        c.number_format = _FORMATO_MONETARIO
        c.fill = PatternFill("solid", fgColor=cor_fundo)
        c.alignment = Alignment(horizontal="right")
        c.border = _borda_fina()
        _celula(ws, i, 6, g.get("acao", "")[:150], cor_fundo, wrap=True)
        ws.row_dimensions[i].height = 30

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


# ============================================================
# ABA GATILHOS DE HISTORICO
# ============================================================

def _criar_aba_gatilhos_historico(wb, gatilhos):
    ws = wb.create_sheet("Gatilhos_Hist")
    historicos = [g for g in gatilhos if g.get("tipo") == "historico"]

    cabecalhos = [
        "Gatilho", "Severidade", "Categoria", "Conta",
        "Valor (R$)", "Data", "Historico (trecho)", "Acao"
    ]
    larguras = [10, 12, 20, 30, 18, 12, 50, 50]
    _cabecalho_tabela(ws, 1, cabecalhos, larguras)

    ordem_sev = {"alta": 0, "media": 1, "baixa": 2}
    hist_ord = sorted(historicos, key=lambda x: (ordem_sev.get(x.get("severidade", "baixa"), 3), -x.get("valor", 0)))

    for i, g in enumerate(hist_ord, start=2):
        sev = g.get("severidade", "baixa")
        cor_sev = _COR_ALTA if sev == "alta" else (_COR_MEDIA if sev == "media" else _COR_BAIXA)
        linha_par = i % 2 == 0
        cor_fundo = _COR_ALTERNADA if linha_par else _COR_BRANCO

        _celula(ws, i, 1, g.get("gatilho", ""), cor_fundo, negrito=True, centralizado=True)
        _celula(ws, i, 2, sev.upper(), cor_sev, cor_texto="FFFFFF", centralizado=True, negrito=True)
        _celula(ws, i, 3, g.get("categoria", ""), cor_fundo, centralizado=True)
        _celula(ws, i, 4, g.get("conta", ""), cor_fundo)
        c = ws.cell(i, 5, g.get("valor", 0))
        c.number_format = _FORMATO_MONETARIO
        c.fill = PatternFill("solid", fgColor=cor_fundo)
        c.alignment = Alignment(horizontal="right")
        c.border = _borda_fina()
        _celula(ws, i, 6, str(g.get("data", ""))[:10], cor_fundo, centralizado=True)
        hist_txt = str(g.get("historico", ""))[:120]
        _celula(ws, i, 7, hist_txt, cor_fundo, wrap=True)
        _celula(ws, i, 8, g.get("acao", "")[:120], cor_fundo, wrap=True)
        ws.row_dimensions[i].height = 35

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"


# ============================================================
# UTILITARIOS
# ============================================================

def _cabecalho_tabela(ws, linha: int, cabecalhos: list[str], larguras: list[int]):
    """Insere linha de cabecalho com formatacao padrao."""
    for col, (cab, larg) in enumerate(zip(cabecalhos, larguras), start=1):
        c = ws.cell(linha, col, cab)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_COR_CABEC)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _borda_fina()
        ws.column_dimensions[get_column_letter(col)].width = larg
    ws.row_dimensions[linha].height = 22


def _celula(
    ws, linha: int, col: int, valor,
    cor_fundo: str = _COR_BRANCO,
    cor_texto: str = "000000",
    negrito: bool = False,
    centralizado: bool = False,
    wrap: bool = False,
    span: int = 1,
):
    """Cria celula formatada."""
    c = ws.cell(linha, col, valor)
    c.font = Font(bold=negrito, color=cor_texto)
    c.fill = PatternFill("solid", fgColor=cor_fundo)
    c.alignment = Alignment(
        horizontal="center" if centralizado else "left",
        vertical="center",
        wrap_text=wrap,
    )
    c.border = _borda_fina()
    return c


def _mesclar_titulo(ws, intervalo: str, texto: str, cor: str, tamanho: int = 11):
    """Mescla celulas e aplica estilo de titulo."""
    ws.merge_cells(intervalo)
    inicio = intervalo.split(":")[0]
    col = 1
    for i, c in enumerate(inicio):
        if c.isdigit():
            linha = int(inicio[i:])
            break
        col = ord(c.upper()) - ord("A") + 1

    c = ws.cell(linha, col, texto)
    c.font = Font(bold=True, color="FFFFFF", size=tamanho)
    c.fill = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _borda_fina()
